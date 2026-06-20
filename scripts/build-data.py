#!/usr/bin/env python3
"""Construye los archivos de datos de la quiniela.

Fuentes (en scripts/sources/):
  - quiniela_pdf.txt : el PDF original convertido con `pdftotext -layout` (14 participantes).
  - *.xlsx           : quinielas adicionales en Excel (un participante por archivo).
Combinado con el dataset openfootball (fechas/sedes/marcadores reales).

Genera: data/fixtures.js, data/participants.js, data/results.js
Reejecutar tras agregar/editar una quiniela.  No es necesario para el sitio en runtime.

Requiere: openpyxl (solo si hay archivos .xlsx)."""
import re, json, sys, unicodedata, datetime, os, glob

HERE = os.path.dirname(__file__)
SOURCES = os.path.join(HERE, "sources")
PDF_TXT = sys.argv[1] if len(sys.argv) > 1 else os.path.join(SOURCES, "quiniela_pdf.txt")
OF_JSON = sys.argv[2] if len(sys.argv) > 2 else "/tmp/of2026.json"
OUT = os.path.join(HERE, "..", "data")

# Nombre del participante por archivo Excel (la celda del nombre suele venir vacía)
EXCEL_NAMES = {
    "Quiniela_FamRico_2026_Lorena_Fabian.xlsx": "LORENA Y FABIÁN",
}

# --- Registro de equipos: clave canonica EN (openfootball) -> display ES + bandera
TEAMS = {
    "Germany":("Alemania","🇩🇪"), "Saudi Arabia":("Arabia Saudita","🇸🇦"),
    "Algeria":("Argelia","🇩🇿"), "Argentina":("Argentina","🇦🇷"),
    "Australia":("Australia","🇦🇺"), "Austria":("Austria","🇦🇹"),
    "Bosnia & Herzegovina":("Bosnia y Herzegovina","🇧🇦"), "Brazil":("Brasil","🇧🇷"),
    "Belgium":("Bélgica","🇧🇪"), "Cape Verde":("Cabo Verde","🇨🇻"),
    "Canada":("Canadá","🇨🇦"), "Qatar":("Catar","🇶🇦"), "Colombia":("Colombia","🇨🇴"),
    "South Korea":("Corea del Sur","🇰🇷"), "Ivory Coast":("Costa de Marfil","🇨🇮"),
    "Croatia":("Croacia","🇭🇷"), "Curaçao":("Curazao","🇨🇼"), "Ecuador":("Ecuador","🇪🇨"),
    "Egypt":("Egipto","🇪🇬"), "Scotland":("Escocia","🏴󠁧󠁢󠁳󠁣󠁴󠁿"), "Spain":("España","🇪🇸"),
    "USA":("Estados Unidos","🇺🇸"), "France":("Francia","🇫🇷"), "Ghana":("Ghana","🇬🇭"),
    "Haiti":("Haití","🇭🇹"), "England":("Inglaterra","🏴󠁧󠁢󠁥󠁮󠁧󠁿"), "Iraq":("Irak","🇮🇶"),
    "Iran":("Irán","🇮🇷"), "Japan":("Japón","🇯🇵"), "Jordan":("Jordania","🇯🇴"),
    "Morocco":("Marruecos","🇲🇦"), "Mexico":("México","🇲🇽"), "Norway":("Noruega","🇳🇴"),
    "New Zealand":("Nueva Zelanda","🇳🇿"), "Netherlands":("Países Bajos","🇳🇱"),
    "Panama":("Panamá","🇵🇦"), "Paraguay":("Paraguay","🇵🇾"), "Portugal":("Portugal","🇵🇹"),
    "DR Congo":("RD Congo","🇨🇩"), "Czech Republic":("República Checa","🇨🇿"),
    "Senegal":("Senegal","🇸🇳"), "South Africa":("Sudáfrica","🇿🇦"),
    "Sweden":("Suecia","🇸🇪"), "Switzerland":("Suiza","🇨🇭"), "Tunisia":("Túnez","🇹🇳"),
    "Turkey":("Turquía","🇹🇷"), "Uruguay":("Uruguay","🇺🇾"), "Uzbekistan":("Uzbekistán","🇺🇿"),
}
# Texto del PDF (normalizado, incluye typos) -> clave EN canonica
ES2EN = {
    "ALEMANIA":"Germany","ARABIA SAUDITA":"Saudi Arabia","ARGELIA":"Algeria",
    "ARGENTINA":"Argentina","AUSTRALIA":"Australia","AUSTRIA":"Austria",
    "BOSNIA HERZEGOVINA":"Bosnia & Herzegovina","BRASIL":"Brazil","BÉLGICA":"Belgium",
    "CABO VERDE":"Cape Verde","CANADÁ":"Canada","CATAR":"Qatar","COLOMBIA":"Colombia",
    "COREA":"South Korea","KOREA":"South Korea","COSTA DE MARFIL":"Ivory Coast",
    "CROACIA":"Croatia","CURAZAO":"Curaçao","ECUADOR":"Ecuador","EGIPTO":"Egypt",
    "ESCOCIA":"Scotland","ESPAÑA":"Spain","ESTADOS UNIDOS":"USA","FRANCIA":"France",
    "GHANA":"Ghana","HAITI":"Haiti","INGLATERRA":"England","IRAK":"Iraq","IRÁN":"Iran",
    "JAPÓN":"Japan","JORDANIA":"Jordan","MARRUECOS":"Morocco","MÉXICO":"Mexico",
    "NORUEGA":"Norway","NUEVA ZELANDA":"New Zealand","PAISES BAJOS":"Netherlands",
    "PAISES BJOS":"Netherlands","PANAMÁ":"Panama","PARAGUAY":"Paraguay",
    "PORTUGAL":"Portugal","POTUGAL":"Portugal","RD CONGO":"DR Congo",
    "REP. CHECA":"Czech Republic","SENEGAL":"Senegal","SUDÁFRICA":"South Africa",
    "SUECIA":"Sweden","SUIZA":"Switzerland","TUNEZ":"Tunisia","TÚNEZ":"Tunisia",
    "TURQUIA":"Turkey","URUGUAY":"Uruguay","UZBEKISTÁN":"Uzbekistan",
}

def slug(s):
    s = unicodedata.normalize("NFKD", s).encode("ascii","ignore").decode()
    return re.sub(r"[^a-z0-9]+","-", s.lower()).strip("-")

def mid(group_letter, home_en, away_en):
    return f"{group_letter}-{slug(home_en)}-vs-{slug(away_en)}"

# --- Parseo del PDF -----------------------------------------------------------
with open(PDF_TXT) as f:
    lines=[l.rstrip("\n") for l in f]
blocks=[]; cur=None
for l in lines:
    if "PARTICIPANTE" in l and "RESULTADO" in l:
        if cur: blocks.append(cur)
        name=l.split("PARTICIPANTE",1)[1].split("RESULTADO",1)[0].strip()
        cur={"name":name,"lines":[]}
    elif cur is not None:
        cur["lines"].append(l)
if cur: blocks.append(cur)

def parse_line(l):
    if len(l)<86 or l[24:25] not in "ABCDEFGHIJKL": return None
    g=l[24]; partido=l[25:64].strip()
    if not re.search(r".\s+[-\.]\s+.", partido): return None
    picks=[k for col,k in [(67,"L"),(76,"E"),(84,"V")] if "X" in l[col-3:col+4]]
    pick=picks[0] if len(picks)==1 else None
    sc=[]; rest=[]
    for t in l[88:].split():
        if len(sc)<2 and re.fullmatch(r"[0-9Oo]+",t): sc.append(int(t.replace("O","0").replace("o","0")))
        else: rest.append(t)
    ph,pa=(sc+[None,None])[:2]
    a,b=re.split(r"\s+[-\.]\s+",partido,maxsplit=1)
    return g, a.strip().upper(), b.strip().upper(), pick, ph, pa

def en(name_es):
    key=name_es.replace(".","").strip().upper()
    key=re.sub(r"\s+"," ",key)
    # exact then with dot variants
    for k in (name_es.strip().upper(), key):
        if k in ES2EN: return ES2EN[k]
    # try restoring REP. CHECA
    if key.startswith("REP"): return "Czech Republic"
    raise KeyError(name_es)

# fixtures canonicos a partir del primer participante (orden/orientacion del PDF)
ref=[]
for l in blocks[0]["lines"]:
    p=parse_line(l)
    if p: ref.append(p)
assert len(ref)==72, f"esperaba 72 fixtures, obtuve {len(ref)}"

# --- openfootball: resultados/fechas reales -----------------------------------
of=json.load(open(OF_JSON))
of_group=[m for m in of["matches"] if str(m.get("group","")).startswith("Group")]
of_index={}      # frozenset(home,away) -> match
team_group={}    # equipo EN -> letra de grupo REAL (autoritativa)
for m in of_group:
    of_index[frozenset((m["team1"],m["team2"]))]=m
    L=m["group"].replace("Group ","").strip()
    team_group[m["team1"]]=L; team_group[m["team2"]]=L

# El PDF a veces trae mal la letra del grupo (p.ej. la 3a jornada de J/K/L
# aparece como G/H/I). Tomamos el grupo real por membresia de equipos.
def real_group(home_en, away_en):
    g1=team_group.get(home_en); g2=team_group.get(away_en)
    assert g1==g2, f"equipos de distinto grupo: {home_en} ({g1}) vs {away_en} ({g2})"
    return g1

fixtures=[]; results={}
unmatched=[]; relabeled=0
for g_pdf,a_es,b_es,_,_,_ in ref:
    ha=en(a_es); aw=en(b_es)
    key=frozenset((ha,aw))
    m=of_index.get(key)
    g=real_group(ha,aw)
    if g!=g_pdf: relabeled+=1
    i=mid(g,ha,aw)
    fx={"id":i,"group":g,
        "homeEN":ha,"awayEN":aw,
        "homeES":TEAMS[ha][0],"awayES":TEAMS[aw][0],
        "homeFlag":TEAMS[ha][1],"awayFlag":TEAMS[aw][1]}
    if m:
        fx["date"]=m.get("date"); fx["time"]=m.get("time"); fx["ground"]=m.get("ground")
        ft=(m.get("score") or {}).get("ft")
        if ft:
            # orientar al home/away de la quiniela
            if m["team1"]==ha: h,aa=ft[0],ft[1]
            else:              h,aa=ft[1],ft[0]
            results[i]={"h":h,"a":aa,"status":"FT"}
    else:
        unmatched.append((g,ha,aw))
    fixtures.append(fx)

if unmatched:
    print("SIN EMPAREJAR:",unmatched); sys.exit(1)
print(f"Grupos corregidos respecto al PDF: {relabeled}")

# --- participantes ------------------------------------------------------------
participants=[]
for b in blocks:
    preds={}
    pl=[parse_line(l) for l in b["lines"]]; pl=[p for p in pl if p]
    assert len(pl)==72, f"{b['name']}: {len(pl)} predicciones"
    for g_pdf,a_es,b_es,pick,ph,pa in pl:
        ha=en(a_es); aw=en(b_es)
        i=mid(real_group(ha,aw),ha,aw)
        preds[i]={"pick":pick,"ph":ph,"pa":pa}
    participants.append({"name":b["name"],"preds":preds})

# --- participantes en Excel (.xlsx, uno por archivo) --------------------------
# El Excel sigue el mismo orden de calendario que los fixtures, así que mapeamos
# por POSICIÓN (fila i -> partido i). Esto corrige automáticamente la orientación
# local/visitante y tolera errores de captura en el nombre del partido.
SWAP={"L":"V","V":"L","E":"E"}
def parse_excel(path, fixtures):
    import openpyxl
    wb=openpyxl.load_workbook(path, data_only=True)
    raw=[]
    for ws in wb.worksheets:
        for r in range(1, ws.max_row+1):
            g=ws.cell(row=r,column=3).value           # C: grupo
            partido=ws.cell(row=r,column=4).value      # D: "EQUIPO - EQUIPO"
            if not g or not partido: continue
            if not re.fullmatch(r"[A-L]", str(g).strip()): continue
            if " - " not in partido and " . " not in partido: continue
            cells=[ws.cell(row=r,column=c).value for c in (5,6,7)]   # E/F/G = L/E/V
            picks=[k for k,v in zip("LEV",cells) if v and str(v).strip().upper()=="X"]
            pick=picks[0] if len(picks)==1 else None
            nums=re.findall(r"\d+", str(ws.cell(row=r,column=9).value or ""))  # I: "2-0"
            ph,pa=(int(nums[0]),int(nums[1])) if len(nums)>=2 else (None,None)
            a_es,b_es=re.split(r"\s+[-\.]\s+", partido.strip(), maxsplit=1)
            raw.append((a_es,b_es,pick,ph,pa))
    assert len(raw)==72, f"{os.path.basename(path)}: {len(raw)} predicciones (esperaba 72)"
    preds={}; mism=[]
    for i,(a_es,b_es,pick,ph,pa) in enumerate(raw):
        f=fixtures[i]
        ha=en(a_es)
        if frozenset((ha,en(b_es)))!=frozenset((f["homeEN"],f["awayEN"])):
            mism.append((i,f"{a_es} - {b_es}",f["id"]))
        if ha==f["awayEN"]:                 # el Excel trae local/visitante invertido
            pick=SWAP.get(pick,pick); ph,pa=pa,ph
        preds[f["id"]]={"pick":pick,"ph":ph,"pa":pa}
    assert 72-len(mism)>=70, f"{os.path.basename(path)}: el orden no alinea ({len(mism)} desajustes)"
    return preds, mism

for xlsx in sorted(glob.glob(os.path.join(SOURCES,"*.xlsx"))):
    fname=os.path.basename(xlsx)
    name=EXCEL_NAMES.get(fname) or re.sub(r"\.xlsx$","",fname).split("_")[-1].upper()
    preds,mism=parse_excel(xlsx, fixtures)
    participants.append({"name":name,"preds":preds})
    print(f"Excel agregado: {name} <- {fname}")
    for i,txt,fid in mism:
        print(f"   nota: fila {i} del Excel decía '{txt}'; se asignó al partido {fid} por posición")

# --- escribir archivos --------------------------------------------------------
os.makedirs(OUT,exist_ok=True)
def write_js(path,var,obj):
    with open(path,"w") as f:
        f.write(f"window.{var} = ")
        json.dump(obj,f,ensure_ascii=False,indent=1)
        f.write(";\n")

write_js(os.path.join(OUT,"fixtures.js"),"FIXTURES",fixtures)
write_js(os.path.join(OUT,"participants.js"),"PARTICIPANTS",participants)
write_js(os.path.join(OUT,"results.js"),"RESULTS",
         {"updated":datetime.datetime.utcnow().replace(microsecond=0).isoformat()+"Z",
          "source":"openfootball/worldcup.json (snapshot inicial)",
          "matches":results})

# resumen
played=len(results)
print(f"OK  fixtures={len(fixtures)}  participantes={len(participants)}  con_resultado={played}/72")
print("Ejemplo fixture:",json.dumps(fixtures[0],ensure_ascii=False))
